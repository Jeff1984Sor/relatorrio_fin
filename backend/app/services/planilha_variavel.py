"""Planilha do relatório de variável, com as colunas do modelo do escritório."""

from __future__ import annotations

import datetime as dt
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .planilha import COR_ESCURA, FORMATO_MOEDA, _preenchimento
from .variavel import LinhaVariavel, ResumoVariavel

FORMATO_DATA = "DD/MM/YYYY"
FORMATO_PERCENTUAL = "0.0%"

COR_ZEBRA = "F5F7FA"
COR_BORDA_SUAVE = "D9D9D9"
COR_DESTAQUE = "FFF2CC"

BORDA_INFERIOR = Border(bottom=Side(style="hair", color=COR_BORDA_SUAVE))

# (título, atributo, largura, formato). O conjunto e a ordem espelham o modelo
# `Variavel Novo.xlsx` — é o que o escritório já usa para conferir.
COLUNAS: list[tuple[str, str, int, str | None]] = [
    ("Grupo", "grupo", 16, None),
    ("Pagador", "pagador", 40, None),
    ("Cliente", "cliente", 40, None),
    ("NH", "nh", 11, None),
    ("NF", "nf", 12, None),
    ("Situação", "situacao", 11, None),
    ("Data de Vencimento", "data_vencimento", 13, FORMATO_DATA),
    ("Data de Pagamento", "data_pagamento", 13, FORMATO_DATA),
    ("Numero do Caso", "numero_do_caso", 12, "0"),
    ("Título", "titulo", 46, None),
    ("Área", "area", 18, None),
    ("Responsável", "responsavel", 30, None),
    ("Valor Bruto", "valor_bruto", 14, FORMATO_MOEDA),
    ("Valor Pago", "valor_pago", 14, FORMATO_MOEDA),
    ("Imposto", "aliquota", 9, FORMATO_PERCENTUAL),
    ("Valor dos Impostos", "valor_dos_impostos", 14, FORMATO_MOEDA),
    ("Valor Líquido", "valor_liquido", 15, FORMATO_MOEDA),
    ("Participação", "participacao", 11, FORMATO_PERCENTUAL),
    ("Variável", "variavel", 14, FORMATO_MOEDA),
]

LINHA_CABECALHO = 4
COLUNAS_DESTAQUE = {"Variável"}

# Colunas derivadas saem como fórmula, não como número congelado: assim, mexer
# na alíquota ou na participação dentro do Excel recalcula a variável na hora.
FORMULAS = {
    "Valor dos Impostos": "={pago}{linha}*{imposto}{linha}",
    "Valor Líquido": "={pago}{linha}-{impostos}{linha}",
    "Variável": "={liquido}{linha}*{participacao}{linha}",
}

TOTALIZADAS = ("Valor Bruto", "Valor Pago", "Valor dos Impostos", "Valor Líquido", "Variável")


def _letras() -> dict[str, str]:
    """Título da coluna → letra no Excel, para montar as fórmulas."""
    return {
        titulo: get_column_letter(indice)
        for indice, (titulo, *_resto) in enumerate(COLUNAS, start=1)
    }


def _formula(titulo: str, linha: int) -> str:
    letras = _letras()
    return FORMULAS[titulo].format(
        linha=linha,
        pago=letras["Valor Pago"],
        imposto=letras["Imposto"],
        impostos=letras["Valor dos Impostos"],
        liquido=letras["Valor Líquido"],
        participacao=letras["Participação"],
    )


def _titulo(ws: Worksheet, resumo: ResumoVariavel) -> None:
    letra_final = get_column_letter(len(COLUNAS))

    ws.merge_cells(f"A1:{letra_final}1")
    ws["A1"] = "RELATÓRIO DE REMUNERAÇÃO VARIÁVEL"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    periodo = "Período não identificado"
    if resumo.periodo_inicio and resumo.periodo_fim:
        periodo = (
            f"Pagamentos de {resumo.periodo_inicio.strftime('%d/%m/%Y')} "
            f"a {resumo.periodo_fim.strftime('%d/%m/%Y')}"
        )

    ws.merge_cells(f"A2:{letra_final}2")
    ws["A2"] = (
        f"{periodo}  •  Imposto {resumo.aliquota * 100:.1f}%  •  "
        f"{len(resumo.linhas)} linhas  •  Gerado em "
        f"{dt.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    ws["A2"].font = Font(size=9, color="808080")


def _cabecalho(ws: Worksheet) -> None:
    for coluna, (titulo, _campo, largura, _formato) in enumerate(COLUNAS, start=1):
        celula = ws.cell(row=LINHA_CABECALHO, column=coluna, value=titulo)
        celula.font = Font(bold=True, color="FFFFFF", size=10)
        celula.fill = _preenchimento(COR_ESCURA)
        celula.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws.column_dimensions[get_column_letter(coluna)].width = largura

    ws.row_dimensions[LINHA_CABECALHO].height = 30
    ws.freeze_panes = ws.cell(row=LINHA_CABECALHO + 1, column=5)
    ws.auto_filter.ref = f"A{LINHA_CABECALHO}:{get_column_letter(len(COLUNAS))}{LINHA_CABECALHO}"


def _escrever_linha(ws: Worksheet, indice: int, linha: LinhaVariavel, zebra: bool) -> None:
    for coluna, (titulo, campo, _largura, formato) in enumerate(COLUNAS, start=1):
        valor = _formula(titulo, indice) if titulo in FORMULAS else getattr(linha, campo)
        celula = ws.cell(row=indice, column=coluna, value=valor)

        if formato:
            celula.number_format = formato
        celula.border = BORDA_INFERIOR
        celula.alignment = Alignment(
            horizontal="right" if formato else "left",
            vertical="center",
            shrink_to_fit=formato is None,
        )
        if titulo in COLUNAS_DESTAQUE:
            celula.font = Font(bold=True)
            celula.fill = _preenchimento(COR_DESTAQUE)
        elif zebra:
            celula.fill = _preenchimento(COR_ZEBRA)


def _rodape(ws: Worksheet, indice: int, primeira: int) -> None:
    for coluna in range(1, len(COLUNAS) + 1):
        celula = ws.cell(row=indice, column=coluna)
        celula.fill = _preenchimento(COR_ESCURA)
        celula.font = Font(bold=True, color="FFFFFF")
        celula.alignment = Alignment(horizontal="right", vertical="center")

    ws.cell(row=indice, column=1, value="TOTAL").alignment = Alignment(horizontal="left")

    if indice <= primeira:  # relatório sem nenhuma linha
        return

    for coluna, (titulo, _campo, _largura, _formato) in enumerate(COLUNAS, start=1):
        if titulo in TOTALIZADAS:
            letra = get_column_letter(coluna)
            celula = ws.cell(
                row=indice, column=coluna, value=f"=SUM({letra}{primeira}:{letra}{indice - 1})"
            )
            celula.number_format = FORMATO_MOEDA


def _aba_conferencia(ws: Worksheet, resumo: ResumoVariavel) -> None:
    for coluna, texto in enumerate(("Tipo", "O que verificar", "Quantidade", "Detalhes"), 1):
        celula = ws.cell(row=1, column=coluna, value=texto)
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = _preenchimento(COR_ESCURA)

    for indice, aviso in enumerate(resumo.avisos, start=2):
        ws.cell(row=indice, column=1, value=aviso.tipo)
        ws.cell(row=indice, column=2, value=aviso.mensagem).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        ws.cell(row=indice, column=3, value=aviso.quantidade)
        ws.cell(
            row=indice,
            column=4,
            value=", ".join(str(d) for d in aviso.detalhes) if aviso.detalhes else "",
        )

    ws.freeze_panes = "A2"
    for coluna, largura in enumerate((26, 70, 12, 70), start=1):
        ws.column_dimensions[get_column_letter(coluna)].width = largura


def gerar_xlsx(resumo: ResumoVariavel) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório Variável"

    _titulo(ws, resumo)
    _cabecalho(ws)

    primeira = LINHA_CABECALHO + 1
    indice = primeira
    for posicao, linha in enumerate(resumo.linhas):
        _escrever_linha(ws, indice, linha, zebra=posicao % 2 == 1)
        indice += 1

    _rodape(ws, indice, primeira)

    if resumo.avisos:
        _aba_conferencia(wb.create_sheet("Conferência"), resumo)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
