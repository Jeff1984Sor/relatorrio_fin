"""Geração da planilha consolidada com openpyxl."""

from __future__ import annotations

import datetime as dt
import io
import re
import unicodedata
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .agregacao import Resumo

FORMATO_MOEDA = '"R$" #,##0.00_);[Red]("R$" #,##0.00)'
FORMATO_PERCENTUAL = "0.0%"

COR_ESCURA = "404040"
COR_CATEGORIA = "F8CBAD"
COR_BORDA = "BFBFBF"
COR_BORDA_SUAVE = "D9D9D9"

BORDA_FINA = Border(*(Side(style="thin", color=COR_BORDA),) * 4)
BORDA_INFERIOR_HAIR = Border(bottom=Side(style="hair", color=COR_BORDA_SUAVE))

LARGURA_NOME = 46
LARGURA_CONTA = 18
LARGURA_TOTAL = 18
LARGURA_PERCENTUAL = 10
LARGURA_QTD = 12


def _preenchimento(cor: str) -> PatternFill:
    return PatternFill(start_color=cor, end_color=cor, fill_type="solid")


def nome_arquivo_saida(nome_original: str) -> str:
    base = re.sub(r"\.[^.]+$", "", nome_original) or "despesas"
    base = unicodedata.normalize("NFKD", base)
    base = "".join(c for c in base if not unicodedata.combining(c))
    base = re.sub(r"[^A-Za-z0-9._-]+", "-", base).strip("-").lower()
    return f"resumo-{base or 'despesas'}.xlsx"


def _contas_em_coluna(resumo: Resumo) -> list[str]:
    """Uma coluna por conta só faz sentido quando há mais de uma."""
    return resumo.contas if len(resumo.contas) > 1 else []


def _valores_da_linha(
    total: Decimal,
    por_conta: dict[str, Decimal],
    contas: list[str],
    percentual: float,
    qtd: int,
) -> list[object]:
    return [
        *(por_conta.get(conta, Decimal("0")) for conta in contas),
        total,
        percentual,
        qtd,
    ]


def _formatar_numeros(ws: Worksheet, linha: int, total_colunas: int) -> None:
    """Layout: nome | contas… | Total | % | Lançamentos.

    Moeda vai da coluna 2 até a de Total; o percentual é a penúltima; a última
    (contagem) fica com o formato geral.
    """
    for coluna in range(2, total_colunas - 1):
        ws.cell(row=linha, column=coluna).number_format = FORMATO_MOEDA
    ws.cell(row=linha, column=total_colunas - 1).number_format = FORMATO_PERCENTUAL


def _aba_resumo(ws: Worksheet, resumo: Resumo, contexto: dict) -> None:
    ws.title = "Resumo"

    # `summaryBelow = False` põe o total ACIMA do grupo; sem isso o `+/−` da
    # lateral do Excel aparece na linha errada.
    ws.sheet_properties.outlinePr.summaryBelow = False

    contas = _contas_em_coluna(resumo)
    cabecalhos = ["Categoria / Subcategoria", *contas, "Total", "%", "Lançamentos"]
    ultima_coluna = len(cabecalhos)
    letra_final = get_column_letter(ultima_coluna)

    ws.merge_cells(f"A1:{letra_final}1")
    titulo = ws["A1"]
    titulo.value = "RESUMO DE DESPESAS POR CATEGORIA"
    titulo.font = Font(bold=True, size=14)
    titulo.alignment = Alignment(horizontal="left", vertical="center")

    gerado_em = contexto.get("gerado_em") or dt.datetime.now()
    ws.merge_cells(f"A2:{letra_final}2")
    subtitulo = ws["A2"]
    subtitulo.value = (
        f"{_texto_periodo(resumo.periodo_inicio, resumo.periodo_fim)}  •  "
        f"Origem: {contexto.get('nome_arquivo', '-')}  •  "
        f"Gerado em {gerado_em.strftime('%d/%m/%Y %H:%M')}"
    )
    subtitulo.font = Font(size=9, color="808080")

    for coluna, texto in enumerate(cabecalhos, start=1):
        celula = ws.cell(row=4, column=coluna, value=texto)
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = _preenchimento(COR_ESCURA)
        celula.alignment = Alignment(
            horizontal="left" if coluna == 1 else "right", vertical="center", wrap_text=True
        )
    ws.freeze_panes = "A5"

    linha = 5
    for categoria in resumo.categorias:
        _escrever_linha(
            ws,
            linha,
            categoria.rotulo.upper(),
            _valores_da_linha(
                categoria.total, categoria.por_conta, contas, categoria.percentual, categoria.qtd
            ),
            destaque=True,
        )
        linha += 1
        for sub in categoria.subcategorias:
            _escrever_linha(
                ws,
                linha,
                sub.rotulo,
                _valores_da_linha(sub.total, sub.por_conta, contas, sub.percentual, sub.qtd),
                destaque=False,
            )
            ws.row_dimensions[linha].outline_level = 1
            linha += 1

    _escrever_linha(
        ws,
        linha,
        "TOTAL GERAL",
        _valores_da_linha(
            resumo.total_geral,
            resumo.total_por_conta,
            contas,
            1.0 if resumo.total_geral else 0.0,
            resumo.qtd_lancamentos,
        ),
        total=True,
    )

    larguras = [
        LARGURA_NOME,
        *([LARGURA_CONTA] * len(contas)),
        LARGURA_TOTAL,
        LARGURA_PERCENTUAL,
        LARGURA_QTD,
    ]
    for indice, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(indice)].width = largura


def _escrever_linha(
    ws: Worksheet,
    linha: int,
    nome: str,
    valores: list[object],
    destaque: bool = False,
    total: bool = False,
) -> None:
    celula_nome = ws.cell(row=linha, column=1, value=nome)
    if total or destaque:
        celula_nome.font = Font(bold=True, color="FFFFFF" if total else "000000")
    celula_nome.alignment = Alignment(horizontal="left", indent=0 if (destaque or total) else 2)

    for offset, valor in enumerate(valores):
        celula = ws.cell(row=linha, column=2 + offset, value=valor)
        celula.alignment = Alignment(horizontal="right")
        if total or destaque:
            celula.font = Font(bold=True, color="FFFFFF" if total else "000000")

    for coluna in range(1, len(valores) + 2):
        celula = ws.cell(row=linha, column=coluna)
        if total:
            celula.fill = _preenchimento(COR_ESCURA)
        elif destaque:
            celula.fill = _preenchimento(COR_CATEGORIA)
            celula.border = BORDA_FINA
        else:
            celula.border = BORDA_INFERIOR_HAIR

    _formatar_numeros(ws, linha, len(valores) + 1)


def _aba_detalhado(ws: Worksheet, resumo: Resumo) -> None:
    cabecalhos = [
        "Data",
        "Conta",
        "Fornecedor",
        "Categoria",
        "Subcategoria",
        "Valor",
        "Arquivo de origem",
        "Linha na origem",
    ]
    for coluna, texto in enumerate(cabecalhos, start=1):
        celula = ws.cell(row=1, column=coluna, value=texto)
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = _preenchimento(COR_ESCURA)

    for indice, detalhe in enumerate(resumo.detalhado, start=2):
        ws.cell(row=indice, column=1, value=detalhe.data).number_format = "DD/MM/YYYY"
        ws.cell(row=indice, column=2, value=detalhe.conta)
        ws.cell(row=indice, column=3, value=detalhe.fornecedor)
        ws.cell(row=indice, column=4, value=detalhe.categoria)
        ws.cell(row=indice, column=5, value=detalhe.subcategoria)
        ws.cell(row=indice, column=6, value=detalhe.valor).number_format = FORMATO_MOEDA
        ws.cell(row=indice, column=7, value=detalhe.arquivo)
        ws.cell(row=indice, column=8, value=detalhe.linha_origem)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{max(1, len(resumo.detalhado) + 1)}"
    for coluna, largura in enumerate((14, 22, 38, 30, 30, 18, 28, 16), start=1):
        ws.column_dimensions[get_column_letter(coluna)].width = largura


def _aba_conferencia(ws: Worksheet, resumo: Resumo) -> None:
    for coluna, texto in enumerate(("Tipo", "O que verificar", "Quantidade", "Detalhes"), 1):
        celula = ws.cell(row=1, column=coluna, value=texto)
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = _preenchimento(COR_ESCURA)

    for indice, aviso in enumerate(resumo.avisos, start=2):
        ws.cell(row=indice, column=1, value=aviso.tipo)
        ws.cell(row=indice, column=2, value=aviso.mensagem).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        ws.cell(row=indice, column=3, value=aviso.quantidade)
        ws.cell(
            row=indice,
            column=4,
            value=", ".join(str(d) for d in aviso.detalhes) if aviso.detalhes else "",
        )

    ws.freeze_panes = "A2"
    for coluna, largura in enumerate((32, 60, 12, 60), start=1):
        ws.column_dimensions[get_column_letter(coluna)].width = largura


def _texto_periodo(inicio: dt.date | None, fim: dt.date | None) -> str:
    if inicio and fim:
        return f"Período: {inicio.strftime('%d/%m/%Y')} a {fim.strftime('%d/%m/%Y')}"
    return "Período não identificado"


def gerar_xlsx(resumo: Resumo, contexto: dict) -> bytes:
    wb = Workbook()
    _aba_resumo(wb.active, resumo, contexto)
    _aba_detalhado(wb.create_sheet("Detalhado"), resumo)
    if resumo.avisos:
        _aba_conferencia(wb.create_sheet("Conferência"), resumo)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
