import io
from decimal import Decimal

from openpyxl import load_workbook

from app.services.agregacao import extrair_linhas, montar_resumo
from app.services.inspecao import inspecionar
from app.services.leitura import ler_linhas
from app.services.planilha import gerar_xlsx, nome_arquivo_saida
from tests.fixture_builder import TOTAL_ESPERADO


def _resumo(conteudo: bytes):
    linhas, _aba = ler_linhas(conteudo, "base.xlsx")
    cabecalho = inspecionar(linhas)
    detalhes, avisos = extrair_linhas(linhas, cabecalho.indice, cabecalho.mapeamento, True)
    return montar_resumo(detalhes, avisos)


def _abrir(conteudo_xlsx: bytes):
    return load_workbook(io.BytesIO(conteudo_xlsx))


def _coluna(ws, titulo: str) -> int:
    """Acha a coluna pelo cabeçalho — a posição muda com o número de contas."""
    for indice in range(1, ws.max_column + 1):
        if ws.cell(row=4, column=indice).value == titulo:
            return indice
    raise AssertionError(f"coluna {titulo!r} não encontrada")


def test_abas_geradas(base_suja):
    wb = _abrir(gerar_xlsx(_resumo(base_suja), {"nome_arquivo": "base.xlsx"}))
    assert wb.sheetnames == ["Resumo", "Detalhado", "Conferência"]


def test_conferencia_so_existe_quando_ha_avisos(base_sem_subcategoria):
    wb = _abrir(gerar_xlsx(_resumo(base_sem_subcategoria), {"nome_arquivo": "limpa.xlsx"}))
    assert "Conferência" not in wb.sheetnames


def test_summary_below_falso_para_o_mais_menos_ficar_no_lugar_certo(base_suja):
    wb = _abrir(gerar_xlsx(_resumo(base_suja), {"nome_arquivo": "base.xlsx"}))
    assert wb["Resumo"].sheet_properties.outlinePr.summaryBelow is False


def test_subcategorias_ficam_no_nivel_1_do_agrupamento(base_suja):
    wb = _abrir(gerar_xlsx(_resumo(base_suja), {"nome_arquivo": "base.xlsx"}))
    ws = wb["Resumo"]
    niveis = {
        ws.cell(row=linha, column=1).value: ws.row_dimensions[linha].outline_level
        for linha in range(5, ws.max_row)
    }
    assert niveis["DESPESA COM PESSOAL"] == 0
    assert niveis["Água e Luz"] == 1


def test_total_geral_na_ultima_linha_bate_com_a_base(base_suja):
    wb = _abrir(gerar_xlsx(_resumo(base_suja), {"nome_arquivo": "base.xlsx"}))
    ws = wb["Resumo"]
    total = _coluna(ws, "Total")
    assert ws.cell(row=ws.max_row, column=1).value == "TOTAL GERAL"
    assert Decimal(str(ws.cell(row=ws.max_row, column=total).value)) == Decimal(TOTAL_ESPERADO)


def test_colunas_por_conta_somam_o_total_da_linha(base_suja):
    """A base suja tem três contas financeiras: cada uma vira uma coluna."""
    wb = _abrir(gerar_xlsx(_resumo(base_suja), {"nome_arquivo": "base.xlsx"}))
    ws = wb["Resumo"]

    contas = ["Bradesco", "Caixa", "Itaú"]
    for conta in contas:
        _coluna(ws, conta)

    total = _coluna(ws, "Total")
    soma_contas = sum(
        Decimal(str(ws.cell(row=ws.max_row, column=_coluna(ws, c)).value)) for c in contas
    )
    assert soma_contas == Decimal(str(ws.cell(row=ws.max_row, column=total).value))


def test_cabecalho_congelado_e_formatos(base_suja):
    wb = _abrir(gerar_xlsx(_resumo(base_suja), {"nome_arquivo": "base.xlsx"}))
    ws = wb["Resumo"]
    assert ws.freeze_panes == "A5"
    assert ws.cell(row=5, column=_coluna(ws, "Total")).number_format.startswith('"R$"')
    assert ws.cell(row=5, column=_coluna(ws, "%")).number_format == "0.0%"
    assert ws["A4"].value == "Categoria / Subcategoria"


def test_detalhado_preserva_todos_os_lancamentos(base_suja):
    resumo = _resumo(base_suja)
    wb = _abrir(gerar_xlsx(resumo, {"nome_arquivo": "base.xlsx"}))
    ws = wb["Detalhado"]
    assert ws.max_row == resumo.qtd_lancamentos + 1
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None


def test_nome_do_arquivo_de_saida():
    assert nome_arquivo_saida("Relatório Analítico.xlsx") == "resumo-relatorio-analitico.xlsx"
    assert nome_arquivo_saida("base.csv") == "resumo-base.xlsx"
