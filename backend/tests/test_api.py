import io
from decimal import Decimal

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture()
def cliente():
    with TestClient(app) as cliente:
        yield cliente


def _upload(*arquivos: tuple[bytes, str]) -> list[tuple[str, tuple[str, bytes, str]]]:
    """Multipart com o mesmo campo repetido — é assim que vários arquivos sobem."""
    return [("arquivos", (nome, conteudo, XLSX)) for conteudo, nome in arquivos]


def test_inspecionar_devolve_mapeamento_e_amostra(cliente, base_suja):
    resposta = cliente.post(
        "/api/despesas/inspecionar", files=_upload((base_suja, "base-suja.xlsx"))
    )
    assert resposta.status_code == 200

    arquivos = resposta.json()["arquivos"]
    assert len(arquivos) == 1
    assert arquivos[0]["nome"] == "base-suja.xlsx"
    assert arquivos[0]["linha_cabecalho"] == 3
    assert arquivos[0]["colunas"][7] == "Valor Líquido"
    assert arquivos[0]["mapeamento"]["valor"] == 7
    assert len(arquivos[0]["amostra"]) == 10


def test_inspecionar_varios_arquivos(cliente, base_suja, fluxo_de_caixa):
    resposta = cliente.post(
        "/api/despesas/inspecionar",
        files=_upload((base_suja, "base.xlsx"), (fluxo_de_caixa, "fluxo.xlsx")),
    )
    assert resposta.status_code == 200

    arquivos = resposta.json()["arquivos"]
    assert [a["nome"] for a in arquivos] == ["base.xlsx", "fluxo.xlsx"]
    assert arquivos[1]["mapeamento"]["somente_preenchidos"] is True


def test_processar_devolve_o_resumo(cliente, base_suja):
    resposta = cliente.post("/api/despesas/processar", files=_upload((base_suja, "base.xlsx")))
    assert resposta.status_code == 200

    corpo = resposta.json()
    assert Decimal(str(corpo["total_geral"])) == Decimal("-11625.85")
    assert corpo["qtd_lancamentos"] == 11
    assert corpo["arquivos"] == ["base.xlsx"]
    assert corpo["avisos"]
    assert corpo["categorias"]


def test_consolidado_de_varios_arquivos_soma_tudo(cliente, fluxo_de_caixa, fluxo_conta_unica):
    sozinho = cliente.post(
        "/api/despesas/processar", files=_upload((fluxo_de_caixa, "junho.xlsx"))
    ).json()

    juntos = cliente.post(
        "/api/despesas/processar",
        files=_upload((fluxo_de_caixa, "junho.xlsx"), (fluxo_conta_unica, "julho.xlsx")),
    ).json()

    assert juntos["arquivos"] == ["julho.xlsx", "junho.xlsx"]
    assert juntos["qtd_lancamentos"] == sozinho["qtd_lancamentos"] * 2
    assert Decimal(str(juntos["total_geral"])) == Decimal(str(sozinho["total_geral"])) * 2

    # A conta única do segundo arquivo soma no Itaú; o Bradesco vem só do primeiro.
    assert juntos["contas"] == ["Banco Bradesco", "Banco Itaú"]
    assert Decimal(str(juntos["total_por_conta"]["Banco Bradesco"])) == Decimal("-220")


def test_baixar_xlsx_com_coluna_por_conta(cliente, fluxo_de_caixa):
    resposta = cliente.post("/api/despesas/xlsx", files=_upload((fluxo_de_caixa, "junho.xlsx")))
    assert resposta.status_code == 200
    assert "resumo-junho.xlsx" in resposta.headers["content-disposition"]

    ws = load_workbook(io.BytesIO(resposta.content))["Resumo"]
    cabecalhos = [ws.cell(row=4, column=c).value for c in range(1, 7)]
    assert cabecalhos == [
        "Categoria / Subcategoria",
        "Banco Bradesco",
        "Banco Itaú",
        "Total",
        "%",
        "Lançamentos",
    ]

    assert ws.cell(row=ws.max_row, column=1).value == "TOTAL GERAL"
    assert Decimal(str(ws.cell(row=ws.max_row, column=4).value)) == Decimal("-53660.93")


def test_nome_do_download_com_varios_arquivos(cliente, fluxo_de_caixa, fluxo_conta_unica):
    resposta = cliente.post(
        "/api/despesas/xlsx",
        files=_upload((fluxo_de_caixa, "junho.xlsx"), (fluxo_conta_unica, "julho.xlsx")),
    )
    assert resposta.status_code == 200
    assert "resumo-despesas-consolidado.xlsx" in resposta.headers["content-disposition"]


def test_mesmo_arquivo_pode_subir_quantas_vezes_quiser(cliente, base_suja):
    """Não há trava de duplicado nem estado: o resultado é sempre idêntico."""
    respostas = [
        cliente.post("/api/despesas/processar", files=_upload((base_suja, "base.xlsx")))
        for _ in range(5)
    ]
    assert [r.status_code for r in respostas] == [200] * 5
    assert len({r.text for r in respostas}) == 1


def test_arquivo_corrompido_diz_qual_arquivo_falhou(cliente, base_suja):
    resposta = cliente.post(
        "/api/despesas/processar",
        files=_upload((base_suja, "boa.xlsx"), (b"isto nao e uma planilha", "quebrada.xlsx")),
    )
    assert resposta.status_code == 422
    detalhe = resposta.json()["detail"]
    assert "Traceback" not in detalhe
    assert detalhe.startswith("quebrada.xlsx:")


def test_extensao_nao_aceita(cliente):
    resposta = cliente.post("/api/despesas/processar", files=_upload((b"abc", "foto.png")))
    assert resposta.status_code == 422
    assert "não é aceito" in resposta.json()["detail"]


def test_limite_de_arquivos(cliente, base_suja):
    resposta = cliente.post(
        "/api/despesas/processar", files=_upload(*[(base_suja, f"m{i}.xlsx") for i in range(25)])
    )
    assert resposta.status_code == 422
    assert "até 24 arquivos" in resposta.json()["detail"]


def test_mapeamento_manual_sobrescreve_a_deteccao(cliente, base_suja):
    resposta = cliente.post(
        "/api/despesas/processar",
        files=_upload((base_suja, "base.xlsx")),
        data={"mapeamento": '{"valor": 4, "categoria": 8, "subcategoria": 9, "data": 0}'},
    )
    assert resposta.status_code == 200
    assert Decimal(str(resposta.json()["total_geral"])) == Decimal("-11625.85")


def test_mapeamento_invalido(cliente, base_suja):
    resposta = cliente.post(
        "/api/despesas/processar",
        files=_upload((base_suja, "base.xlsx")),
        data={"mapeamento": "isto nao e json"},
    )
    assert resposta.status_code == 422
    assert "mapeamento" in resposta.json()["detail"].lower()


def test_categoria_concatenada_pela_api(cliente, base_categoria_concatenada):
    resposta = cliente.post(
        "/api/despesas/processar", files=_upload((base_categoria_concatenada, "junho.xlsx"))
    )
    assert resposta.status_code == 200

    rotulos = [c["rotulo"] for c in resposta.json()["categorias"]]
    assert rotulos == ["CERTIDÕES", "DESPESA COM PESSOAL", "DESPESAS ADMINISTRATIVAS"]


def test_rotas_de_estado_nao_existem_mais(cliente):
    assert cliente.get("/api/despesas").status_code == 404
    assert cliente.get("/api/despesas/1").status_code == 404


def test_saude(cliente):
    assert cliente.get("/api/saude").json() == {"status": "ok"}
