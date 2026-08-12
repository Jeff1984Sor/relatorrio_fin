import io
from decimal import Decimal

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app
from app.services import spreadsheetml, variavel
from app.services.leitura import ler_linhas
from tests import fixture_variavel

XLS_MIME = "application/vnd.ms-excel"


@pytest.fixture(scope="session")
def cubo(tmp_path_factory) -> bytes:
    destino = tmp_path_factory.mktemp("variavel") / "VisaoCuboLocal.xls"
    return fixture_variavel.construir_cubo(destino).read_bytes()


@pytest.fixture(scope="session")
def casos(tmp_path_factory) -> bytes:
    destino = tmp_path_factory.mktemp("variavel") / "Relatorio Casos.xlsx"
    return fixture_variavel.construir_casos(destino).read_bytes()


@pytest.fixture()
def resumo(cubo, casos):
    linhas_cubo, _ = ler_linhas(cubo, "VisaoCuboLocal.xls")
    linhas_casos, _ = ler_linhas(casos, "Relatorio Casos.xlsx")
    return variavel.montar(linhas_cubo, linhas_casos)


# --- leitura do formato XML ---------------------------------------------------


def test_reconhece_o_xml_disfarcado_de_xls(cubo):
    assert spreadsheetml.parece_spreadsheetml(cubo)


def test_mescla_vertical_vira_valor_repetido(cubo):
    """O cubo mescla o recebimento e lista um caso por linha — o parser preenche."""
    linhas, aba = ler_linhas(cubo, "VisaoCuboLocal.xls")
    assert aba == "Page 1"

    # Nº 5065 ocupa duas linhas: a segunda só traz o caso, o resto vem da mescla.
    assert linhas[1][0] == "Nº 5065"
    assert linhas[2][0] == "Nº 5065"
    assert linhas[2][9] == "valor: 343.26" or linhas[1][9] == "valor: 343.26"
    assert linhas[1][10] == "1945"
    assert linhas[2][10] == "1949"


@pytest.mark.parametrize(
    "bruto,esperado",
    [
        ("valor: 3000.00", Decimal("3000.00")),
        ("valor: 2.815,50", Decimal("2815.50")),
        ("VALOR: 180.00", Decimal("180.00")),
        ("(Blank)", None),
        ("", None),
        (None, None),
        (1234.5, Decimal("1234.5")),
    ],
)
def test_parse_do_valor_do_cubo(bruto, esperado):
    assert variavel.parse_valor_cubo(bruto) == esperado


# --- cálculo ------------------------------------------------------------------


def test_formula_da_variavel(resumo):
    """Impostos sobre o pago, líquido = pago - impostos, variável = líquido x participação."""
    linha = next(l for l in resumo.linhas if l.nh == "Nº 5072")

    assert linha.valor_pago == Decimal("2111.62")
    assert linha.valor_dos_impostos == Decimal("369.53")
    assert linha.valor_liquido == Decimal("1742.09")
    assert linha.participacao == Decimal("0.3")
    assert linha.variavel == Decimal("522.63")
    assert linha.responsavel == "Alice Rocha Assuncao"
    assert linha.area == "Cível"


def test_area_do_caso_entra_no_relatorio(resumo):
    areas = {l.nh: l.area for l in resumo.linhas}
    assert areas["Nº 5065"] == "Imobiliário"
    # Recebimento sem caso não tem área.
    assert areas["Nº 5040"] == ""


def test_relatorio_de_casos_sem_a_coluna_area_continua_funcionando(cubo, tmp_path):
    """A Área é opcional: quem exportar sem ela ainda consegue gerar o relatório."""
    from openpyxl import Workbook

    destino = tmp_path / "casos-sem-area.xlsx"
    wb = Workbook()
    ws = wb.active
    for coluna, titulo in enumerate(
        ("Número do Caso", "Título", "Responsável", "Participação"), start=1
    ):
        ws.cell(row=1, column=coluna, value=titulo)
    ws.append([2299, "MA Hipódromo", "Alice Rocha Assuncao", 0.3])
    wb.save(destino)

    linhas_cubo, _ = ler_linhas(cubo, "cubo.xls")
    linhas_casos, _ = ler_linhas(destino.read_bytes(), "casos.xlsx")
    resumo = variavel.montar(linhas_cubo, linhas_casos)

    linha = next(l for l in resumo.linhas if l.nh == "Nº 5072")
    assert linha.area == ""
    assert linha.responsavel == "Alice Rocha Assuncao"


def test_nh_nao_repete_para_o_mesmo_responsavel(resumo):
    """Dois casos do mesmo responsável saem numa linha só, com o valor inteiro."""
    linhas = [l for l in resumo.linhas if l.nh == "Nº 5065"]

    assert len(linhas) == 1
    assert linhas[0].valor_pago == Decimal("343.26")
    assert linhas[0].casos_do_responsavel == 2


def test_rateio_proporcional_entre_responsaveis(resumo):
    """3 casos, 2 responsáveis (2 de um, 1 do outro): 2/3 e 1/3 do valor."""
    linhas = [l for l in resumo.linhas if l.nh == "Nº 5066"]
    assert len(linhas) == 2

    por_responsavel = {l.responsavel: l for l in linhas}
    rubens = por_responsavel["Rubens Leonardo Marin"]
    luiz = por_responsavel["Luiz Roberto Hijo Sampietro"]

    assert rubens.casos_do_responsavel == 2
    assert luiz.casos_do_responsavel == 1
    assert rubens.valor_pago == Decimal("343.26")
    assert luiz.valor_pago == Decimal("171.63")
    # Nenhum centavo criado nem perdido no rateio.
    assert rubens.valor_pago + luiz.valor_pago == Decimal("514.89")


def test_rateio_nunca_perde_centavo():
    assert sum(variavel._ratear(Decimal("100.00"), [1, 1, 1])) == Decimal("100.00")
    assert sum(variavel._ratear(Decimal("0.05"), [1, 1, 1, 1, 1, 1])) == Decimal("0.05")
    assert sum(variavel._ratear(Decimal("5574.69"), [29, 4])) == Decimal("5574.69")
    # O centavo que sobra vai para quem tem mais casos.
    assert variavel._ratear(Decimal("1.00"), [2, 1]) == [Decimal("0.67"), Decimal("0.33")]


def test_recebimento_sem_caso_entra_sem_responsavel(resumo):
    linha = next(l for l in resumo.linhas if l.nh == "Nº 5040")

    assert linha.responsavel == ""
    assert linha.numero_do_caso is None
    assert linha.variavel == Decimal("0")
    # O valor continua na tabela, para o total bater com o cubo.
    assert linha.valor_pago == Decimal("2815.50")


def test_caso_inexistente_no_relatorio_de_casos(resumo):
    linha = next(l for l in resumo.linhas if l.nh == "Nº 5090")
    assert linha.responsavel == ""
    assert linha.variavel == Decimal("0")

    aviso = next(a for a in resumo.avisos if a.tipo == "caso_nao_encontrado")
    assert 999999 in aviso.detalhes


def test_total_pago_bate_com_o_cubo(resumo):
    """Invariante principal: o rateio não pode alterar o total recebido."""
    esperado = sum(
        Decimal(r[9]) for r in fixture_variavel.RECEBIMENTOS
    )
    assert resumo.total_pago == esperado


def test_total_por_responsavel_bate_com_a_soma_das_linhas(resumo):
    soma_linhas = sum(l.variavel for l in resumo.linhas)
    assert sum(resumo.por_responsavel.values()) == soma_linhas
    assert resumo.total_variavel == soma_linhas


def test_aliquota_configuravel(cubo, casos):
    linhas_cubo, _ = ler_linhas(cubo, "cubo.xls")
    linhas_casos, _ = ler_linhas(casos, "casos.xlsx")
    resumo = variavel.montar(linhas_cubo, linhas_casos, aliquota=Decimal("0"))

    linha = next(l for l in resumo.linhas if l.nh == "Nº 5072")
    assert linha.valor_dos_impostos == Decimal("0")
    assert linha.valor_liquido == linha.valor_pago


def test_periodo_vem_da_data_de_pagamento(resumo):
    assert resumo.periodo_inicio.isoformat() == "2026-07-13"
    assert resumo.periodo_fim.isoformat() == "2026-07-30"


# --- API ----------------------------------------------------------------------


@pytest.fixture()
def cliente():
    with TestClient(app) as cliente:
        yield cliente


def _envio(cubo: bytes, casos: bytes):
    return {
        "cubo": ("VisaoCuboLocal.xls", cubo, XLS_MIME),
        "casos": ("Relatorio Casos.xlsx", casos, XLS_MIME),
    }


def test_processar_pela_api(cliente, cubo, casos):
    resposta = cliente.post("/api/variavel/processar", files=_envio(cubo, casos))
    assert resposta.status_code == 200

    corpo = resposta.json()
    assert Decimal(str(corpo["aliquota"])) == Decimal("0.175")
    assert corpo["linhas"]
    assert corpo["por_responsavel"]["Alice Rocha Assuncao"] == "522.63"


def test_aliquota_aceita_virgula_e_percentual(cliente, cubo, casos):
    for entrada in ("17,5", "0,175", "17.5"):
        resposta = cliente.post(
            "/api/variavel/processar", files=_envio(cubo, casos), data={"aliquota": entrada}
        )
        assert resposta.status_code == 200
        assert Decimal(str(resposta.json()["aliquota"])) == Decimal("0.175")


def test_aliquota_invalida(cliente, cubo, casos):
    resposta = cliente.post(
        "/api/variavel/processar", files=_envio(cubo, casos), data={"aliquota": "abc"}
    )
    assert resposta.status_code == 422
    assert "alíquota" in resposta.json()["detail"].lower()


def test_xlsx_do_variavel(cliente, cubo, casos):
    resposta = cliente.post("/api/variavel/xlsx", files=_envio(cubo, casos))
    assert resposta.status_code == 200
    assert "relatorio-variavel" in resposta.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(resposta.content))
    ws = wb["Relatório Variável"]

    from app.services.planilha_variavel import COLUNAS

    cabecalhos = [ws.cell(row=4, column=c).value for c in range(1, len(COLUNAS) + 1)]
    assert cabecalhos == [titulo for titulo, *_ in COLUNAS]
    assert cabecalhos[0] == "Grupo"
    assert cabecalhos[3] == "NH"
    assert "Área" in cabecalhos
    assert cabecalhos[-1] == "Variável"

    assert ws.cell(row=ws.max_row, column=1).value == "TOTAL"
    assert "Conferência" in wb.sheetnames


def test_arquivo_de_casos_sem_as_colunas_certas(cliente, cubo):
    resposta = cliente.post(
        "/api/variavel/processar",
        files={
            "cubo": ("cubo.xls", cubo, XLS_MIME),
            "casos": ("errado.csv", b"a;b;c\n1;2;3\n", "text/csv"),
        },
    )
    assert resposta.status_code == 422
    assert "casos" in resposta.json()["detail"].lower()
