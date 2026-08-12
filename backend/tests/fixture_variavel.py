"""Fixtures do relatório de variável, reproduzindo os formatos reais.

O cubo sai do sistema como SpreadsheetML 2003 (XML com extensão `.xls`), sem
cabeçalho, com valores em texto (`valor: 3000.00`), `(Blank)` no lugar de vazio
e células mescladas verticalmente quando um recebimento cobre vários casos.
"""

from __future__ import annotations

from pathlib import Path
from xml.sax.saxutils import escape

from openpyxl import Workbook

NS = 'xmlns="urn:schemas-microsoft-com:office:spreadsheet" ' 'xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet"'

# (NH, pagador, cliente, grupo, NF, situação, bruto, vencimento, pagamento, pago, [casos])
RECEBIMENTOS = [
    # Um caso só, o comum.
    ("Nº 5072", "MA Hipódromo SPE Ltda.", "Mundo Apto Ltda", "Mundo Apto", "Sem Vínculo",
     "Quitada", "2250.00", "10/07/2026", "13/07/2026", "2111.62", ["2299"]),
    # Dois casos do MESMO responsável: uma linha só, valor inteiro.
    ("Nº 5065", "MA Sarzedas SPE Ltda", "Mundo Apto Ltda", "Mundo Apto", "Sem Vínculo",
     "Quitada", "360.00", "10/07/2026", "13/07/2026", "343.26", ["1945", "1949"]),
    # Três casos de DOIS responsáveis: rateio proporcional, duas linhas.
    ("Nº 5066", "MA Hipódromo SPE Ltda.", "Mundo Apto Ltda", "Mundo Apto", "Sem Vínculo",
     "Quitada", "540.00", "10/07/2026", "13/07/2026", "514.89", ["1947", "1948", "1950"]),
    # Sem caso vinculado: entra na tabela, sem responsável e sem variável.
    ("Nº 5040", "JJN Participações LTDA", "JJN Participações LTDA", "(Blank)", "Nº 6605",
     "Quitada", "3000.00", "30/05/2026", "21/07/2026", "2815.50", ["(Blank)"]),
    # Caso que não existe no relatório de casos.
    ("Nº 5090", "MA Itaberaba Ltda", "MA Itaberaba Ltda", "Mundo Apto", "Nº 6569",
     "Quitada", "180.00", "31/07/2026", "30/07/2026", "180.00", ["999999"]),
]

# (número, título, área, responsável, participação)
CASOS = [
    (1945, "Regularização Sarzedas", "Imobiliário", "Rubens Leonardo Marin", 0.3),
    (1947, "Diligência Hipódromo I", "Imobiliário", "Rubens Leonardo Marin", 0.3),
    (1948, "Diligência Hipódromo II", "Imobiliário", "Rubens Leonardo Marin", 0.3),
    (1949, "Regularização Sarzedas II", "Imobiliário", "Rubens Leonardo Marin", 0.3),
    (1950, "Diligência Hipódromo III", "Cível", "Luiz Roberto Hijo Sampietro", 0.3),
    (2299, "MA Hipódromo - Defesa - Multa sobre Ruídos", "Cível", "Alice Rocha Assuncao", 0.3),
]


def _celula(valor: str, mescla_vertical: int = 0) -> str:
    atributo = f' ss:MergeDown="{mescla_vertical}"' if mescla_vertical else ""
    return f'<Cell{atributo}><Data ss:Type="String">{escape(valor)}</Data></Cell>'


def construir_cubo(caminho: Path) -> Path:
    """Escreve o cubo em SpreadsheetML, com as mesclas verticais de verdade."""
    linhas_xml = []
    for recebimento in RECEBIMENTOS:
        *comuns, casos = recebimento
        mescla = len(casos) - 1
        # O cubo exporta as colunas de dinheiro como texto: `valor: 2111.62`.
        comuns = [
            f"valor: {v}" if posicao in (6, 9) else str(v)
            for posicao, v in enumerate(comuns)
        ]

        primeira = "".join(_celula(v, mescla) for v in comuns)
        primeira += _celula(casos[0])
        linhas_xml.append(f"<Row>{primeira}</Row>")

        # As linhas seguintes só trazem o caso; o resto vem da mescla de cima.
        for caso in casos[1:]:
            linhas_xml.append(f'<Row><Cell ss:Index="11">'
                              f'<Data ss:Type="String">{escape(caso)}</Data></Cell></Row>')

    xml = (
        '<?xml version="1.0"?>\r\n'
        '<?mso-application progid="Excel.Sheet"?>\r\n'
        f"<Workbook {NS}>"
        '<Worksheet ss:Name="Page 1"><Table>'
        + "".join(linhas_xml)
        + "</Table></Worksheet></Workbook>"
    )
    caminho.parent.mkdir(parents=True, exist_ok=True)
    caminho.write_bytes(xml.encode("utf-8"))
    return caminho


def construir_casos(caminho: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Casos"

    for coluna, titulo in enumerate(
        ("Número do Caso", "Título", "Área", "Responsável", "Participação"), start=1
    ):
        ws.cell(row=1, column=coluna, value=titulo)

    for indice, linha in enumerate(CASOS, start=2):
        for coluna, valor in enumerate(linha, start=1):
            ws.cell(row=indice, column=coluna, value=valor)

    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho)
    return caminho
