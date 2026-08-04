"""Gera a planilha de teste — dados fictícios, com os casos sujos do escopo.

Serve tanto para os testes (via conftest) quanto para inspeção manual:
    python -m tests.fixture_builder tests/fixtures/base-suja.xlsx
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook

CABECALHO = [
    "Data de Pagamento",
    "Data de Vencimento",
    "Conta Financeira",
    "Fornecedor",
    "Valor bruto",
    "Multa + Juros - Desconto",
    "Impostos",
    "Valor Líquido",
    "Categoria",
    "Subcategoria",
]

# (data, conta, fornecedor, bruto, multa, impostos, líquido, categoria, subcategoria)
LINHAS = [
    # As três variantes de "Guias/Custas Judiciais" precisam virar uma linha só.
    ("05/01/2026", "Itaú", "Cartório Central", "(R$ 1.000,00)", 0, 0, "(R$ 1.000,00)",
     "DESPESAS ADMINISTRATIVAS", "Guias / Custas Judiciais"),
    ("06/01/2026", "Itaú", "Cartório Norte", "(R$ 250,50)", 0, 0, "(R$ 250,50)",
     "Despesas Administrativas", "Guias/Custas Judiciais"),
    ("07/01/2026", "Itaú", "Cartório Sul", "-1.234,56", 0, 0, "-1.234,56",
     "DESPESAS ADMINISTRATIVAS", "GUIAS / CUSTAS JUDICIAIS"),

    # Valor numérico puro e valor com R$.
    ("08/01/2026", "Bradesco", "Papelaria Alfa", -300.25, 0, 0, -300.25,
     "DESPESAS ADMINISTRATIVAS", "Material de Escritório"),
    ("09/01/2026", "Bradesco", "Reembolso Cliente X", "R$ 7.901,36", 0, 0, "R$ 7.901,36",
     "DESPESAS REEMBOLSÁVEIS", "Deslocamento"),

    # Linha com valor vazio no meio da base — vira aviso, não entra no total.
    ("10/01/2026", "Bradesco", "Fornecedor Sem Valor", "", 0, 0, "",
     "DESPESAS ADMINISTRATIVAS", "Material de Escritório"),

    # Lançamento sem categoria.
    ("11/01/2026", "Caixa", "Fornecedor Sem Classificação", "(R$ 99,90)", 0, 0,
     "(R$ 99,90)", "", "Diversos"),

    # Subcategoria com ';' — erro de cadastro real na base.
    ("12/01/2026", "Caixa", "BPO Contábil", "(R$ 4.117,00)", 0, 0, "(R$ 4.117,00)",
     "DESPESAS ADMINISTRATIVAS", "BPO Financeiro; DESPESAS ADMINISTRATIVAS"),

    # Acentuação para conferir a ordenação pt-BR (Água antes de Bônus).
    ("13/01/2026", "Caixa", "Concessionária", "(R$ 480,00)", 0, 0, "(R$ 480,00)",
     "DESPESA COM PESSOAL", "Água e Luz"),
    ("14/01/2026", "Caixa", "Equipe", "(R$ 12.000,00)", 0, 0, "(R$ 12.000,00)",
     "DESPESA COM PESSOAL", "Bônus"),

    # Subcategoria vazia.
    ("15/01/2026", "Itaú", "Tarifa Bancária", "(R$ 45,00)", 0, 0, "(R$ 45,00)",
     "DESPESAS FINANCEIRAS", ""),

    # Valor zero, entra na contagem sem mexer no total.
    ("16/01/2026", "Itaú", "Ajuste", "R$ 0,00", 0, 0, "R$ 0,00",
     "DESPESAS FINANCEIRAS", "Ajustes"),
]

# Total esperado da coluna Valor Líquido (a linha vazia fica de fora):
#   -1000.00 -250.50 -1234.56 -300.25 +7901.36 -99.90 -4117.00 -480.00 -12000.00 -45.00 +0.00
TOTAL_ESPERADO = "-11625.85"
QTD_ESPERADA = 11


def construir(caminho: Path, linhas_antes_do_cabecalho: int = 3) -> Path:
    """Escreve a planilha. O cabeçalho fica fora da linha 1 de propósito."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Analítico"

    ws["A1"] = "ESCRITÓRIO MODELO LTDA"
    ws["A2"] = "Relatório analítico de despesas pagas — 01/01/2026 a 31/01/2026"
    # Linha 3 fica em branco; o cabeçalho cai na linha 4.

    for coluna, titulo in enumerate(CABECALHO, start=1):
        ws.cell(row=linhas_antes_do_cabecalho + 1, column=coluna, value=titulo)

    for indice, linha in enumerate(LINHAS, start=linhas_antes_do_cabecalho + 2):
        data, conta, fornecedor, bruto, multa, impostos, liquido, cat, sub = linha
        valores = (data, data, conta, fornecedor, bruto, multa, impostos, liquido, cat, sub)
        for coluna, valor in enumerate(valores, start=1):
            ws.cell(row=indice, column=coluna, value=valor)

    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho)
    return caminho


def construir_sem_subcategoria(caminho: Path) -> Path:
    """Variante sem a coluna Subcategoria — precisa processar mesmo assim."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Analítico"

    colunas = ["Data de Pagamento", "Fornecedor", "Valor Líquido", "Categoria"]
    for coluna, titulo in enumerate(colunas, start=1):
        ws.cell(row=1, column=coluna, value=titulo)

    dados = [
        ("05/01/2026", "Cartório Central", "(R$ 1.000,00)", "DESPESAS ADMINISTRATIVAS"),
        ("06/01/2026", "Equipe", "(R$ 2.000,00)", "DESPESA COM PESSOAL"),
    ]
    for indice, linha in enumerate(dados, start=2):
        for coluna, valor in enumerate(linha, start=1):
            ws.cell(row=indice, column=coluna, value=valor)

    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho)
    return caminho


def construir_categoria_concatenada(caminho: Path) -> Path:
    """Variante como o sistema de gestão exporta de verdade: `CATEGORIA:SUBCATEGORIA`
    numa coluna só, sem coluna de subcategoria."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Analítico"

    colunas = ["Data de Pagamento", "Fornecedor", "Valor Líquido", "Categoria"]
    for coluna, titulo in enumerate(colunas, start=1):
        ws.cell(row=1, column=coluna, value=titulo)

    dados = [
        ("05/06/2026", "Equipe", "(R$ 96.700,00)", "DESPESA COM PESSOAL:REMUNERAÇÃO FIXA"),
        ("06/06/2026", "Equipe", "(R$ 2.761,89)", "DESPESA COM PESSOAL:REMUNERAÇÃO VARIÁVEL"),
        ("07/06/2026", "Plano", "(R$ 300,00)", "DESPESA COM PESSOAL:BENEFÍCIOS"),
        ("08/06/2026", "Revista", "(R$ 455,32)", "DESPESAS ADMINISTRATIVAS:ASSINATURAS"),
        ("09/06/2026", "OAB", "(R$ 735,80)", "DESPESAS ADMINISTRATIVAS:ASSOCIAÇÕES"),
        # Sem `:` — categoria sem subcategoria, tem que continuar funcionando.
        ("10/06/2026", "Cartório", "(R$ 1.624,66)", "CERTIDÕES"),
        # Erro de cadastro real: dois níveis colados com `;` no meio.
        ("11/06/2026", "BPO", "(R$ 4.953,15)",
         "DESPESAS ADMINISTRATIVAS:BPO FINANCEIRO; DESPESAS ADMINISTRATIVAS:CONTABILIDADE"),
    ]
    for indice, linha in enumerate(dados, start=2):
        for coluna, valor in enumerate(linha, start=1):
            ws.cell(row=indice, column=coluna, value=valor)

    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho)
    return caminho


# --- Formato "Fluxo de Caixa" (Movimentos Financeiros) -----------------------
# Cabeçalho na linha 9, coluna 0 vazia, `Categoria / Subcategoria` numa coluna
# só separada por ` : `, e Débito/Crédito em colunas separadas com `-` no lugar
# do valor ausente.

CABECALHO_FLUXO = [
    "",
    "Data",
    "Banco/Conta Financeira",
    "Fornecedor",
    "Categoria / Subcategoria",
    "Observação",
    "Débito",
    "Crédito",
    "Total",
]

# (data, conta, fornecedor, categoria, observação, débito, crédito)
LINHAS_FLUXO = [
    ("01/06/2026", "Banco Itaú", "Intermédica",
     "  Despesas com Pessoal\t\t\t : Plano de saúde", "", -1377.84, "-"),
    ("05/06/2026", "Banco Itaú", "Renata Silva da Costa",
     "  Despesas com Pessoal\t\t\t : Prolabore", "", -1442.69, "-"),
    ("01/06/2026", "Banco Itaú", "Banco Itaú - Taxas",
     "Despesas Operacionais\t\t : Taxas bancárias", "", -126.40, "-"),
    ("10/06/2026", "Banco Itaú", "R Simão Contabilidade",
     "Despesas Operacionais\t\t : Serviços contábeis", "", -270, "-"),
    # Crédito: fica de fora do relatório, sem virar aviso de erro.
    ("01/06/2026", "Banco Itaú", "Tokio Marine", "Honorários Recebidos",
     "Pagamento parcela 1", "-", 12000),
    ("09/06/2026", "Banco Itaú", "Cláudio Oliveira", "Honorários Recebidos",
     "Pagamento parcela 1", "-", 670),
    # Transferência entre contas próprias: entra normalmente, como categoria.
    ("08/06/2026", "Banco Itaú", "Transferência", "Transferência para: Investimentos",
     "", -40000, "-"),
    # Folha sem categoria: só a lista de benefícios, separada por vírgula.
    ("05/06/2026", "Banco Itaú", "Lucas Caramés",
     "Antecipação de dividendos, Auxílio Alimentação, INSS, Prolabore", "", -3621, "-"),
    ("05/06/2026", "Banco Itaú", "Aline Garbin",
     "Antecipação de dividendos, Auxílio Alimentação, INSS, Prolabore", "", -6603, "-"),
    # Segunda conta bancária, para exercitar a coluna por conta.
    ("12/06/2026", "Banco Bradesco", "Cursos para Equipe",
     "Despesas Operacionais\t\t : Cursos e treinamentos", "", -220, "-"),
]

#  -1377.84 -1442.69 -126.40 -270 -40000 -3621 -6603 -220
TOTAL_FLUXO_ESPERADO = "-53660.93"
QTD_FLUXO_ESPERADA = 8


def construir_fluxo_de_caixa(caminho: Path, conta_unica: bool = False) -> Path:
    """Planilha no formato do fluxo de caixa. `conta_unica` deixa tudo no Itaú."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimentos Financeiros"

    ws["A2"] = "Movimentos Financeiros"
    ws["A3"] = "RCA | ESCRITÓRIO MODELO"
    ws["E5"] = "Período: 01/06/2026 à 30/06/2026"
    ws["A6"] = "Saldo Inicial: R$58.370,25"

    for coluna, titulo in enumerate(CABECALHO_FLUXO, start=1):
        ws.cell(row=9, column=coluna, value=titulo or None)

    for indice, linha in enumerate(LINHAS_FLUXO, start=10):
        data, conta, fornecedor, categoria, observacao, debito, credito = linha
        if conta_unica:
            conta = "Banco Itaú"
        valores = (None, data, conta, fornecedor, categoria, observacao, debito, credito, "")
        for coluna, valor in enumerate(valores, start=1):
            if valor not in (None, ""):
                ws.cell(row=indice, column=coluna, value=valor)

    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho)
    return caminho


if __name__ == "__main__":  # pragma: no cover
    destino = Path(sys.argv[1] if len(sys.argv) > 1 else "tests/fixtures/base-suja.xlsx")
    print(construir(destino))
